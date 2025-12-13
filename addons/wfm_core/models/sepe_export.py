from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import timedelta
import base64
import logging

_logger = logging.getLogger(__name__)


class WfmSepeExport(models.Model):
    _name = 'wfm.sepe.export'
    _description = 'SEPE Export Batch'
    _order = 'create_date desc'
    _inherit = ['mail.thread']

    name = fields.Char(
        required=True,
        default=lambda self: _('New'),
        readonly=True
    )

    # Date range
    date_from = fields.Date(required=True, string='From Date')
    date_to = fields.Date(required=True, string='To Date')

    # Visits in this export
    visit_ids = fields.Many2many(
        'wfm.visit',
        'wfm_sepe_export_visit_rel',
        'export_id',
        'visit_id',
        string='Visits',
        domain=[('state', '=', 'done')]
    )
    visit_count = fields.Integer(
        compute='_compute_visit_count',
        string='Visit Count'
    )

    # Workflow state
    state = fields.Selection([
        ('draft', 'Draft'),
        ('exported', 'Exported'),
        ('submitted', 'Submitted to SEPE'),
    ], default='draft', tracking=True, string='Status')

    # Export file
    export_file = fields.Binary(
        string='Excel File',
        attachment=True
    )
    export_filename = fields.Char(string='Filename')

    # Audit fields
    exported_by = fields.Many2one('res.users', readonly=True, string='Exported By')
    export_date = fields.Datetime(readonly=True, string='Export Date')
    submitted_date = fields.Datetime(readonly=True, string='Submitted Date')

    # Statistics
    total_hours = fields.Float(
        compute='_compute_totals',
        string='Total Hours',
        store=True
    )
    total_amount = fields.Monetary(
        compute='_compute_totals',
        string='Total Amount',
        currency_field='currency_id',
        store=True
    )
    currency_id = fields.Many2one(
        'res.currency',
        default=lambda self: self.env.company.currency_id
    )

    @api.depends('visit_ids')
    def _compute_visit_count(self):
        for record in self:
            record.visit_count = len(record.visit_ids)

    @api.depends('visit_ids', 'visit_ids.duration', 'visit_ids.partner_payment_amount')
    def _compute_totals(self):
        for record in self:
            record.total_hours = sum(record.visit_ids.mapped('duration'))
            record.total_amount = sum(record.visit_ids.mapped('partner_payment_amount'))

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', _('New')) == _('New'):
                vals['name'] = self.env['ir.sequence'].next_by_code('wfm.sepe.export') or _('New')
        return super().create(vals_list)

    def action_generate_excel(self):
        """Generate SEPE-compliant Excel export."""
        self.ensure_one()

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise UserError(_('openpyxl library is required. Install with: pip install openpyxl'))

        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SEPE Export'

        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', wrap_text=True)

        # Headers (Greek SEPE format)
        headers = [
            'Α/Α',
            'Ημερομηνία Επίσκεψης',
            'Επωνυμία Πελάτη',
            'ΑΦΜ Πελάτη',
            'Εγκατάσταση',
            'Διεύθυνση',
            'Πόλη',
            'Ώρα Έναρξης',
            'Ώρα Λήξης',
            'Διάρκεια (ώρες)',
            'Τύπος Υπηρεσίας',
            'Ονοματεπώνυμο Συνεργάτη',
            'Ειδικότητα',
            'Τηλέφωνο Συνεργάτη',
            'Κατάσταση',
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Specialty labels in Greek
        specialty_labels = {
            'physician': 'Ιατρός Εργασίας',
            'safety_engineer': 'Τεχνικός Ασφαλείας',
            'health_scientist': 'Επιστήμονας Υγείας',
        }

        # Visit type labels in Greek
        visit_type_labels = {
            'regular': 'Τακτική Επίσκεψη',
            'urgent': 'Επείγουσα',
            'follow_up': 'Επανεξέταση',
        }

        # State labels in Greek
        state_labels = {
            'draft': 'Πρόχειρο',
            'assigned': 'Ανατεθειμένο',
            'confirmed': 'Επιβεβαιωμένο',
            'in_progress': 'Σε Εξέλιξη',
            'done': 'Ολοκληρωμένο',
            'cancelled': 'Ακυρωμένο',
        }

        # Data rows
        for idx, visit in enumerate(self.visit_ids.sorted(key=lambda v: v.visit_date), 1):
            # Format time as HH:MM
            start_hours = int(visit.start_time)
            start_mins = int((visit.start_time % 1) * 60)
            end_hours = int(visit.end_time)
            end_mins = int((visit.end_time % 1) * 60)
            start_time = f"{start_hours:02d}:{start_mins:02d}"
            end_time = f"{end_hours:02d}:{end_mins:02d}"

            ws.append([
                idx,
                visit.visit_date.strftime('%d/%m/%Y') if visit.visit_date else '',
                visit.client_id.name or '',
                visit.client_id.vat or '',
                visit.installation_id.name or '',
                visit.installation_id.address or '',
                visit.installation_id.city or '',
                start_time,
                end_time,
                round(visit.duration, 2),
                visit_type_labels.get(visit.visit_type, visit.visit_type or ''),
                visit.partner_id.name or '',
                specialty_labels.get(visit.partner_id.specialty, visit.partner_id.specialty or ''),
                visit.partner_id.phone or '',
                state_labels.get(visit.state, visit.state or ''),
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to binary
        output = BytesIO()
        wb.save(output)

        filename = f"SEPE_Export_{self.date_from}_{self.date_to}.xlsx"

        self.write({
            'export_file': base64.b64encode(output.getvalue()),
            'export_filename': filename,
            'state': 'exported',
            'export_date': fields.Datetime.now(),
            'exported_by': self.env.user.id,
        })

        # Mark visits as exported
        self.visit_ids.write({
            'sepe_exported': True,
            'sepe_export_date': fields.Datetime.now(),
        })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Export Complete'),
                'message': _('Successfully exported %s visits to Excel.') % len(self.visit_ids),
                'type': 'success',
                'sticky': False,
            }
        }

    def action_submit_to_sepe(self):
        """Mark as submitted to SEPE."""
        self.ensure_one()

        if self.state != 'exported':
            raise UserError(_('You can only submit exported batches.'))

        self.write({
            'state': 'submitted',
            'submitted_date': fields.Datetime.now(),
        })

        # Update billing status of visits to 'invoiced' (ready for invoicing)
        self.visit_ids.write({
            'billing_status': 'invoiced',
        })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Submitted to SEPE'),
                'message': _('Export batch marked as submitted. Visits are now ready for invoicing.'),
                'type': 'success',
            }
        }

    def action_reset_to_draft(self):
        """Reset to draft state."""
        self.ensure_one()
        self.write({
            'state': 'draft',
            'export_file': False,
            'export_filename': False,
            'export_date': False,
            'exported_by': False,
            'submitted_date': False,
        })

    def action_view_visits(self):
        """Open visits in this export."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Exported Visits'),
            'res_model': 'wfm.visit',
            'domain': [('id', 'in', self.visit_ids.ids)],
            'view_mode': 'list,form',
            'target': 'current',
        }

    @api.model
    def _cron_daily_export(self):
        """Cron: Auto-export yesterday's completed visits.

        Called daily by scheduled action.
        """
        yesterday = fields.Date.today() - timedelta(days=1)

        # Find unexported completed visits from yesterday
        visits = self.env['wfm.visit'].search([
            ('visit_date', '=', yesterday),
            ('state', '=', 'done'),
            ('sepe_exported', '=', False),
        ])

        if not visits:
            _logger.info(f"SEPE cron: No visits to export for {yesterday}")
            return True

        _logger.info(f"SEPE cron: Exporting {len(visits)} visits for {yesterday}")

        # Create export batch
        export = self.create({
            'date_from': yesterday,
            'date_to': yesterday,
            'visit_ids': [(6, 0, visits.ids)],
        })

        # Generate Excel
        export.action_generate_excel()

        _logger.info(f"SEPE cron: Successfully created export {export.name}")

        return True
